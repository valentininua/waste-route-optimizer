from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook

from app.services.schedule import parse_frequency_weeks, parse_service_days


@dataclass(frozen=True)
class ParsedPoint:
    original_order: int
    address: str
    bin_code: str | None
    service_day_code: str | None
    service_days: list[str]
    frequency_code: str | None
    frequency_weeks: int | None
    volume: float
    containers: int
    lat: float | None = None
    lng: float | None = None
    service_date: str | None = None
    time_spent: str | None = None


@dataclass(frozen=True)
class ParsedRoute:
    route_code: str
    route_date: str | None
    source_row: int
    declared_total_volume: float | None
    declared_total_containers: int | None
    points: list[ParsedPoint] = field(default_factory=list)


COLUMN_ALIASES = {
    "date": ["date"],
    "time_spent": ["time spent"],
    "order": ["order number", "route order", "order number of th", "order number of the route"],
    "address": ["address"],
    "bin_code": ["code of bin", "bin code"],
    "service_day_code": ["day of week", "service provided"],
    "frequency": ["number of weeks", "weeks when service", "frequency"],
    "volume": ["volume of bin", "volume"],
    "containers": ["number of conta", "niumber of conta", "number of containers", "containers"],
    "lat": ["lat", "latitude"],
    "lng": ["lng", "lon", "longitude"],
}

ROUTE_HEADER_RE = re.compile(r"^ML\s+(?P<number>\d+)\s+no\s+(?P<date>\d{2}\.\d{2}\.\d{2,4})", re.IGNORECASE)
SERVICE_CODE_RE = re.compile(r"^[xX1-7]{7}$")


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _to_float(value: Any, default: float | None = 0) -> float | None:
    text = _clean_cell(value).replace(",", ".")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def _to_int(value: Any, default: int | None = 0) -> int | None:
    number = _to_float(value, None)
    if number is None:
        return default
    try:
        return int(round(number))
    except (ValueError, TypeError):
        return default


def _normalize_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _clean_cell(value)
    if not text:
        return None
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def _route_header(value: Any) -> tuple[str, str | None] | None:
    text = _clean_cell(value)
    match = ROUTE_HEADER_RE.match(text)
    if not match:
        return None
    return f"ML {match.group('number')}", _normalize_date(match.group("date"))


def _route_header_in_row(row: list[Any]) -> tuple[str, str | None] | None:
    for value in row:
        header = _route_header(value)
        if header:
            return header
    return None


def _find_header_row(raw_rows: list[list[Any]]) -> int:
    for idx, row in enumerate(raw_rows[:30], start=1):
        row_text = " ".join(_clean_cell(v).lower() for v in row)
        if "address" in row_text and ("order" in row_text or "route" in row_text):
            return idx
    return 1


def _normalize_columns(header: list[Any]) -> dict[str, int]:
    normalized = [_clean_cell(c).lower() for c in header]
    result: dict[str, int] = {}
    used: set[int] = set()
    for canonical, aliases in COLUMN_ALIASES.items():
        for idx, col in enumerate(normalized):
            if idx in used and canonical == "date":
                continue
            if any(alias in col for alias in aliases):
                # The source file contains two Date columns. The second one is the per-row service date.
                if canonical == "date" and idx == 0 and normalized.count("date") > 1:
                    continue
                result[canonical] = idx
                used.add(idx)
                break
    return result


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _is_total_address(address: str) -> bool:
    lower = address.strip().lower()
    return lower in {"total:", "total", "kopā ml:", "kopa ml:", "kopā:", "kopa:"}


def _looks_like_collection_row(row: list[Any], col: dict[str, int]) -> bool:
    address = _clean_cell(_cell(row, col.get("address")))
    if not address or _is_total_address(address):
        return False
    order = _to_int(_cell(row, col.get("order")), None)
    if order is None or order <= 0:
        return False
    # Avoid accidental parsing of notes or repeated headers.
    if address.lower() == "address":
        return False
    return True


def _build_point(row: list[Any], col: dict[str, int], auto_order: int) -> ParsedPoint:
    order = _to_int(_cell(row, col.get("order")), auto_order) or auto_order
    address = _clean_cell(_cell(row, col.get("address")))
    frequency_code = _clean_cell(_cell(row, col.get("frequency"))) if "frequency" in col else None
    service_code = _clean_cell(_cell(row, col.get("service_day_code"))) if "service_day_code" in col else None
    lat = _to_float(_cell(row, col.get("lat")), None) if "lat" in col else None
    lng = _to_float(_cell(row, col.get("lng")), None) if "lng" in col else None
    return ParsedPoint(
        original_order=order,
        address=address,
        bin_code=_clean_cell(_cell(row, col.get("bin_code"))) if "bin_code" in col else None,
        service_day_code=service_code,
        service_days=parse_service_days(service_code),
        frequency_code=frequency_code,
        frequency_weeks=parse_frequency_weeks(frequency_code),
        volume=float(_to_float(_cell(row, col.get("volume")), 0) or 0) if "volume" in col else 0,
        containers=int(_to_int(_cell(row, col.get("containers")), 0) or 0) if "containers" in col else 0,
        lat=float(lat) if lat is not None else None,
        lng=float(lng) if lng is not None else None,
        service_date=_normalize_date(_cell(row, col.get("date"))) if "date" in col else None,
        time_spent=_clean_cell(_cell(row, col.get("time_spent"))) if "time_spent" in col else None,
    )


def _read_rows(path: str | Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def parse_excel_routes(path: str | Path, include_empty: bool = False) -> list[ParsedRoute]:
    """Parse a source Excel file into separate route blocks.

    The real test file stores many routes in one sheet. Each block starts with a row like
    "ML 11840 no 01.01.26" in the first column and the block total in the Address/Volume/Container columns.
    This parser keeps each ML block separate and skips subtotal/header/service rows.
    """
    rows = _read_rows(path)
    if not rows:
        raise ValueError("Excel file is empty")

    header_row_number = _find_header_row(rows)
    header = rows[header_row_number - 1]
    col = _normalize_columns(header)
    required = {"address", "order"}
    missing = sorted(required - set(col))
    if missing:
        raise ValueError(f"Excel file is missing required columns: {', '.join(missing)}")

    has_route_headers = any(_route_header_in_row(row) for row in rows[header_row_number:])

    routes: list[ParsedRoute] = []
    current_code: str | None = None if has_route_headers else "ROUTE 1"
    current_date: str | None = None
    current_source_row = header_row_number + 1
    declared_volume: float | None = None
    declared_containers: int | None = None
    points: list[ParsedPoint] = []

    def flush() -> None:
        nonlocal points, current_code, current_date, current_source_row, declared_volume, declared_containers
        if current_code is None:
            return
        if points or include_empty:
            routes.append(
                ParsedRoute(
                    route_code=current_code,
                    route_date=current_date,
                    source_row=current_source_row,
                    declared_total_volume=declared_volume,
                    declared_total_containers=declared_containers,
                    points=sorted(points, key=lambda p: p.original_order),
                )
            )
        points = []

    for row_number, row in enumerate(rows[header_row_number:], start=header_row_number + 1):
        header = _route_header_in_row(row)
        if header:
            flush()
            current_code, current_date = header
            current_source_row = row_number
            declared_volume = _to_float(_cell(row, col.get("volume")), None)
            declared_containers = _to_int(_cell(row, col.get("containers")), None)
            continue

        if current_code is None:
            continue
        if not _looks_like_collection_row(row, col):
            continue
        points.append(_build_point(row, col, auto_order=len(points) + 1))

    flush()

    if not routes:
        raise ValueError("No route blocks with collection points found in Excel file")
    return routes

